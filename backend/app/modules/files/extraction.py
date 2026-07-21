import email as email_stdlib
import io

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader

AUDIO_MIME_TYPES = {"audio/mpeg", "audio/wav", "audio/x-wav", "audio/mp4", "audio/x-m4a"}


def extract_text(*, mime_type: str, content: bytes) -> tuple[str | None, str]:
    if mime_type in AUDIO_MIME_TYPES:
        return None, "unsupported"

    try:
        if mime_type == "text/plain":
            return content.decode("utf-8", errors="replace"), "extracted"

        if mime_type == "message/rfc822":
            return _extract_eml_text(content), "extracted"

        if mime_type == "application/pdf":
            reader = PdfReader(io.BytesIO(content))
            text = "\n".join(page.extract_text() or "" for page in reader.pages)
            return text, "extracted"

        if mime_type == (
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        ):
            document = Document(io.BytesIO(content))
            text = "\n".join(paragraph.text for paragraph in document.paragraphs)
            return text, "extracted"

        if mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
            rows = []
            for sheet in workbook.worksheets:
                rows.append(f"Sheet: {sheet.title}")
                for row in sheet.iter_rows(values_only=True):
                    values = [str(value) for value in row if value is not None]
                    if values:
                        rows.append(" | ".join(values))
            return "\n".join(rows), "extracted"
    except Exception:
        return None, "failed"

    return None, "unsupported"


def _extract_eml_text(content: bytes) -> str:
    message = email_stdlib.message_from_bytes(content)
    if not message.is_multipart():
        payload = message.get_payload(decode=True)
        if not isinstance(payload, bytes):
            return ""
        return payload.decode(message.get_content_charset() or "utf-8", errors="replace")

    parts = []
    for part in message.walk():
        if part.get_content_type() == "text/plain":
            payload = part.get_payload(decode=True)
            if isinstance(payload, bytes):
                parts.append(payload.decode(part.get_content_charset() or "utf-8", errors="replace"))
    return "\n".join(parts)
